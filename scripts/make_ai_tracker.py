"""Generates the MyGenie Practical AI content + video tracker (Excel).
Seed outcome numbers are PLACEHOLDER until the owner confirms real figures."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

GREEN, DEEP, ORANGE, SAND, LINE, WHITE = "18A84A", "10402A", "F26822", "F6F8F5", "E6EAE6", "FFFFFF"
thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(bold=True, color=WHITE, size=10)
title_font = Font(bold=True, color=DEEP, size=15)
sub_font = Font(color="5B6B61", size=9, italic=True)
cell_font = Font(size=9, color="14201A")
wrap = Alignment(wrap_text=True, vertical="top")

wb = openpyxl.Workbook()


def make_sheet(ws, title, subtitle, headers, rows, widths):
    ws.sheet_view.showGridLines = False
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    ws.cell(row=1, column=1, value=title).font = title_font
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    ws.cell(row=2, column=1, value=subtitle).font = sub_font
    hr = 4
    for j, h in enumerate(headers, 1):
        c = ws.cell(row=hr, column=j, value=h)
        c.font = hdr_font
        c.fill = PatternFill("solid", fgColor=GREEN)
        c.alignment = wrap
        c.border = border
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = widths[j - 1]
    for r, row in enumerate(rows, hr + 1):
        for j, val in enumerate(row, 1):
            c = ws.cell(row=r, column=j, value=val)
            c.font = cell_font
            c.alignment = wrap
            c.border = border
            if r % 2 == 0:
                c.fill = PatternFill("solid", fgColor=SAND)
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)


HEAD = ["#", "Feature", "URL", "Headline", "Pain", "What the AI does",
        "Seed Outcome (PLACEHOLDER)", "Your Real Metric (fill)", "Device", "Length",
        "Video shot-list (record this exact flow)", "On-screen captions", "Video status", "Poster status"]
W = [4, 20, 16, 26, 34, 38, 24, 22, 14, 8, 50, 34, 12, 12]

ai_rows = [
    [1, "AI Menu Import", "/ai#menu-import", "Go live in minutes, not days",
     "Building a full menu by hand takes days; every typo delays go-live.",
     "Upload your menu and AI auto-builds items, categories, modifiers and prices.",
     "~30 min from file to live POS", "", "Phone (portrait)", "~20s",
     "Tap 'Import Menu' > upload menu PDF/photo > show AI parsing > items/categories/prices auto-populate > tap an item to show it's editable. End on the fully built menu.",
     "Upload your menu | AI builds it | Live in minutes", "Pending", "Pending"],
    [2, "AI Customer Insights", "/ai#customer-insights", "Know who's loyal, who's leaving",
     "You serve hundreds of guests but can't tell who's loyal or slipping.",
     "AI reads every bill to surface regulars, at-risk customers, top spenders, favourites.",
     "30% better repeat rate", "", "Tablet/Desktop", "~20s",
     "Open Customers dashboard > AI insight cards animate (At-risk, Top spenders, Favourite items) > expand one card. End on an at-risk list with a win-back nudge.",
     "AI reads your data | Spots who's slipping | Who to win back", "Pending", "Pending"],
    [3, "Smart Cross-sell & Upsell", "/ai#smart-upsell", "Lift every bill, automatically",
     "Staff forget add-ons, leaving easy revenue on every table.",
     "At billing, AI suggests the right add-on at the right moment.",
     "+18% average bill value", "", "Phone (portrait)", "~15s",
     "In billing/captain flow add a main item (e.g. Pizza) > AI suggested add-on chip appears ('Add Garlic Bread +Rs120') > tap it > bill total rises. End on updated total.",
     "Add an item | AI suggests the upsell | Bigger bills", "Pending", "Pending"],
    [4, "AI Report Audit", "/ai#report-audit", "Numbers you can trust",
     "Reports are only useful if trusted; manual auditing is slow and misses manipulation.",
     "AI auto-audits every report, flagging anomalies, voids, discounts and stock variance.",
     "100% of reports auto-audited", "", "Tablet/Desktop", "~20s",
     "Open any report (sales/discount/inventory) > AI Audit badge shows 'N issues found' > expand > flagged lines (post-payment void, over-policy discount, stock variance) > tap one to investigate.",
     "Every report auto-audited | Anomalies flagged | Numbers you can trust", "Pending", "Pending"],
    [5, "Operational Recommendations", "/ai#ops-recommendations", "What to fix first",
     "Data tells you what happened, not what to do about it.",
     "AI returns prioritised recommendations: what's hurting profit and what to fix first.",
     "12% less wastage", "", "Tablet/Desktop", "~18s",
     "Open 'Recommendations' > prioritised AI cards (high wastage on X, slow-moving item run combo) > tap a card to show the action detail.",
     "AI scans your ops | Flags profit leaks | What to fix first", "Pending", "Pending"],
    [6, "Smart Validations", "/ai#smart-validations", "Protect your margin",
     "Billing errors, over-discounting and post-payment edits drain margin.",
     "AI validates transactions in real time, catching errors and misuse before they cost you.",
     "Rs 1 Lakh leakage caught in 2 weeks (REAL)", "", "Phone (portrait)", "~15s",
     "Apply an oversized discount > AI warning/approval flag pops ('Discount exceeds policy') > show an attempted post-payment edit getting flagged.",
     "Someone bends the rules | AI catches it instantly | Margin protected", "Pending", "Pending"],
    [7, "AI CRM Segmentation", "/ai#crm-segmentation", "Offers that actually land",
     "Blasting the same offer to everyone wastes money and annoys customers.",
     "AI auto-groups customers (regulars, lapsed, big spenders, birthdays) for targeting.",
     "2x campaign conversion", "", "Tablet/Desktop", "~18s",
     "Open CRM > AI auto-segments (Regulars, Lapsed, Big spenders, Birthday this month) > select a segment > push a WhatsApp offer.",
     "AI groups customers | Target the right people | Offers that convert", "Pending", "Pending"],
]

ci_rows = [
    [1, "Central Inventory (PRODUCT PAGE)", "/product/central-inventory", "One stock brain for every outlet",
     "Multi-outlet stock managed in silos causes stock-outs, overstock and variance.",
     "Central visibility + AI auto-reorder + inter-outlet transfers + variance detection across outlets.",
     "20% less stock wastage across outlets", "", "Tablet/Desktop", "~25s (give it the most time)",
     "Central dashboard showing ALL outlets' stock > drill into one outlet's low stock > AI auto-reorder suggestion (demand forecast) > raise an inter-outlet transfer / central indent > show a variance/theft flag where consumption != sales across outlets.",
     "See every outlet's stock live | AI predicts what to reorder | Transfer between outlets in a tap | Catch variance before it costs you", "Pending", "Pending"],
]

specs = [
    ["Device & orientation", "Phone in PORTRAIT (1080x1920) for floor features; Tablet/Desktop (1920x1080) for dashboards, reporting & central inventory."],
    ["Length", "15-25s per clip (Central Inventory up to ~25s). Web autoplay loops must be short."],
    ["Sound", "None. Clips autoplay MUTED + LOOPING, so meaning must read from on-screen caption overlays."],
    ["Captions", "Add 2-4 short caption overlays per clip (see captions column). Keep them large and high-contrast."],
    ["Data hygiene", "Use realistic demo data (plausible dish names, outlets, Rs amounts). NO real customer PII / phone numbers."],
    ["Framing", "Start on the relevant screen already open (skip login/nav). Do the action. END on the AI result / outcome frame."],
    ["Export", "MP4 + WebM, compressed < 2-3 MB each. Also capture ONE hi-res poster screenshot (first meaningful frame) per feature."],
    ["Where files go", "Drop clips in frontend/public/videos/ and posters in frontend/public/video-posters/, then set videoSrc/poster in src/data/ai.js (AI page) or src/data/products.js (Central Inventory)."],
    ["Seed numbers", "All 'Seed Outcome' values are PLACEHOLDERS. Replace with real figures in the 'Your Real Metric' column; main agent will update the site copy."],
]

ws1 = wb.active
ws1.title = "AI Page Content"
make_sheet(ws1, "MyGenie — Practical AI page content & video tracker",
           "Page: /ai  •  All features LIVE in production  •  Seed outcomes are PLACEHOLDER until you confirm real numbers.",
           HEAD, ai_rows, W)

ws2 = wb.create_sheet("Central Inventory")
make_sheet(ws2, "MyGenie — Central Inventory product page & video tracker",
           "Page: /product/central-inventory  •  Franchise / multi-outlet focus.",
           HEAD, ci_rows, W)

ws3 = wb.create_sheet("Recording Specs")
make_sheet(ws3, "Recording specs & checklist (hand to whoever shoots the videos)",
           "Follow these for every clip so they drop straight into the site.",
           ["Topic", "Guideline"], specs, [22, 110])

# ---- FAQ media (selective enrichment across pages) ----
FAQ_HEAD = ["Page", "FAQ question", "Media type", "What to capture (shot / screenshot)", "Inline CTA(s)", "Status"]
FAQ_W = [22, 40, 12, 60, 32, 12]
faq_rows = [
    ["/resources", "What exactly is MyGenie POS?", "Image", "One clean screenshot of the owner dashboard (billing + kitchen + inventory + customers visible).", "See Practical AI; View Pricing", "Pending"],
    ["/resources", "Do I need special hardware?", "Video ~12s", "MyGenie billing running on a normal phone — take an order end to end.", "-", "Pending"],
    ["/resources", "Does it work without stable internet?", "Video ~15s", "Turn off Wi-Fi, bill an order offline, turn Wi-Fi on, show auto-sync confirmation.", "-", "Pending"],
    ["/resources", "Can I manage multiple outlets centrally?", "Image", "Multi-outlet dashboard with several outlets + consolidated numbers.", "Central Inventory; For Chains", "Pending"],
    ["/resources", "How is the AI practical, not hype?", "Video ~15s", "AI Report Audit: open a report, show 'N issues found', expand flagged lines.", "See all 7 AI features", "Pending"],
    ["/ai", "(per-feature clips)", "Video", "Use the 7 shot-lists in the 'AI Page Content' sheet — these also enrich the AI-page FAQ where relevant.", "-", "Pending"],
    ["/product/central-inventory", "Does AI help with reordering?", "Video ~12s", "AI auto-reorder suggestion screen with demand forecast + par-level.", "-", "Pending"],
    ["/product/central-inventory", "Can I transfer stock between outlets?", "Video ~10s", "Raise an inter-outlet transfer with approval + audit trail.", "-", "Pending"],
    ["/solutions/<sector>", "High-impact sector FAQ (1 per sector)", "Image/Video", "Pick the single most decision-critical FAQ per sector and capture the matching screen (e.g. tab management for bars, KOT for QSR).", "-", "To plan after mockup approval"],
    ["/product/<bucket>", "High-impact product FAQ (1 per product)", "Image/Video", "One visual per product page for its most-asked FAQ.", "-", "To plan after mockup approval"],
]
ws4 = wb.create_sheet("FAQ Media")
make_sheet(ws4, "Enriched FAQ media plan (selective — visuals only where they help convert)",
           "FAQs stay on their own pages; these get rich media + inline CTAs. src stays null until you drop assets into public/.",
           FAQ_HEAD, faq_rows, FAQ_W)

wb.save("/app/MyGenie_AI_Content_Tracker.xlsx")
print("Wrote /app/MyGenie_AI_Content_Tracker.xlsx")
