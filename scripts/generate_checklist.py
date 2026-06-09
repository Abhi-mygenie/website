"""
Generate the MyGenie Pre-Production Checklist as a single .xlsx workbook
with one tab per discipline (SEO, GA4, Meta, Google Ads, Deployment) plus a
Summary tab. Source of truth: CR-3B Owner GTM brief, SEO Cutover Gate, and the
CR Control Dashboard.

Run:  python3 /app/scripts/generate_checklist.py
Out:  /app/Pre_Production_Checklist.xlsx
"""
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

OUT = "/app/Pre_Production_Checklist.xlsx"

# ---- styling helpers --------------------------------------------------------
HEADER_FILL = PatternFill("solid", fgColor="1F2A44")        # deep navy
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
TITLE_FONT = Font(bold=True, size=16, color="1F2A44")
SUB_FONT = Font(italic=True, size=9, color="6B7280")
WRAP = Alignment(wrap_text=True, vertical="top")
CENTER = Alignment(horizontal="center", vertical="center", wrap_text=True)
THIN = Side(style="thin", color="D1D5DB")
BORDER = Border(left=THIN, right=THIN, top=THIN, bottom=THIN)

STATUS_FILL = {
    "DONE": PatternFill("solid", fgColor="C6EFCE"),
    "OWNER TODO": PatternFill("solid", fgColor="FFEB9C"),
    "BLOCKED": PatternFill("solid", fgColor="FFC7CE"),
    "READY": PatternFill("solid", fgColor="DDEBF7"),
    "OPTIONAL": PatternFill("solid", fgColor="EDEDED"),
}
PRIORITY_FILL = {
    "P0": PatternFill("solid", fgColor="FFC7CE"),
    "P1": PatternFill("solid", fgColor="FFEB9C"),
    "P2": PatternFill("solid", fgColor="DDEBF7"),
}

COLUMNS = ["#", "Item", "Owner", "Status", "Priority", "Notes"]
WIDTHS = [5, 46, 14, 13, 9, 64]


def style_sheet(ws, title, subtitle, rows):
    # title rows
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(COLUMNS))
    c = ws.cell(row=1, column=1, value=title)
    c.font = TITLE_FONT
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(COLUMNS))
    s = ws.cell(row=2, column=1, value=subtitle)
    s.font = SUB_FONT

    header_row = 4
    for ci, name in enumerate(COLUMNS, start=1):
        cell = ws.cell(row=header_row, column=ci, value=name)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = CENTER
        cell.border = BORDER
        ws.column_dimensions[get_column_letter(ci)].width = WIDTHS[ci - 1]

    for ri, row in enumerate(rows, start=header_row + 1):
        item, owner, status, priority, notes = row
        values = [ri - header_row, item, owner, status, priority, notes]
        for ci, val in enumerate(values, start=1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = BORDER
            cell.alignment = CENTER if ci in (1, 3, 4, 5) else WRAP
        ws.cell(row=ri, column=4).fill = STATUS_FILL.get(status, PatternFill())
        if priority in PRIORITY_FILL:
            ws.cell(row=ri, column=5).fill = PRIORITY_FILL[priority]

    ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
    ws.sheet_view.showGridLines = False


# ---- DATA: each row = (Item, Owner, Status, Priority, Notes) ----------------

SEO = [
    ("Capture Week-0 GSC baseline (indexed pages, top queries, clicks, CTR, position, current 404s)",
     "Owner", "OWNER TODO", "P0", "Before DNS flip. Reference: SEO_Migration_Weekly_Monitoring_Report.md §I."),
    ("Stage 301 redirect layer at Cloudflare (import frontend/cloudflare-bulk-redirects.csv as Bulk Redirect List + rule)",
     "Owner", "READY", "P0", "GATE 1. 19 old→new mappings prepared & validated. Emergent managed hosting alone CANNOT do 301s — must be Cloudflare or nginx."),
    ("Preserve non-www → www 301 (apex mygenie.online → https://www.mygenie.online)",
     "Owner", "READY", "P0", "GATE 2. Already live on Cloudflare — keep the rule active after cutover."),
    ("Confirm canonical host = https://www.mygenie.online across canonical/OG/JSON-LD",
     "Code", "DONE", "P0", "Validated in build JS via SITE_URL. No mixed-host canonical."),
    ("Serve sitemap.xml (48 URLs, www-only) + robots.txt at domain root",
     "Code", "DONE", "P0", "Validated in build artifact. Re-confirm with curl post-cutover (preview shows Cloudflare-managed robots)."),
    ("Verify Google Search Console property for https://www.mygenie.online",
     "Owner", "OWNER TODO", "P0", "DNS TXT or HTML tag. Submit sitemap immediately post-cutover (GATE 4)."),
    ("Facebook domain verification preserved in index.html",
     "Code", "DONE", "P1", "52wnu6dg7xaa18lrghxek38u535rtw present in build."),
    ("Post-cutover QA: all 19 old URLs return real HTTP 301; new + 21 blog pages return 200",
     "Owner", "OWNER TODO", "P0", "Run deploy-time QA block (SEO_Production_Cutover_Gate.md §F). PASS bar before declaring SEO closed."),
    ("Validate OG tags with FB Sharing Debugger + Twitter Card Validator (home + 1 blog post)",
     "Owner", "OWNER TODO", "P1", "After cutover."),
    ("Bing Webmaster — add property + submit sitemap",
     "Owner", "OPTIONAL", "P2", "Optional."),
    ("Prerender/SSR (react-snap) for empty-HTML SPA risk",
     "Code", "OPTIONAL", "P1", "GATE 5 — does not block GO; schedule post-cutover."),
]

GA4 = [
    ("GA4 base tag live via GTM (Measurement ID G-KWHHFEZ5Q3)",
     "Code/Owner", "DONE", "P0", "Loaded through container GTM-K5D84Z3L; host-gated to www.mygenie.online."),
    ("page_view fires on every SPA route change",
     "Code", "DONE", "P0", "Verified via headless dataLayer."),
    ("GA4 - Book Appointment tag on calendly Trigger (Event name demo_booked)",
     "Owner", "DONE", "P0", "Created + container PUBLISHED (13 tags / 9 triggers)."),
    ("Mark demo_booked as a GA4 Key event (after 1 live test booking)",
     "Owner", "OWNER TODO", "P0", "Then import into Google Ads as the 'Book appointment' conversion (account is GA4-linked)."),
    ("Unpause GA4 - OTP Verified tag (trigger lead_verifided)",
     "Owner", "OWNER TODO", "P0", "Currently PAUSED. Trigger name already correct (intentional typo)."),
    ("Map conversion_value (+ currency INR) into GA4 conversion tags",
     "Owner", "OWNER TODO", "P1", "Site sends form_submitted=₹0, lead_verified=₹500, demo_booked=₹2000."),
    ("Register GA4 custom dimensions: otp_verified, form_location, plan_interest (+ optional lead_quality)",
     "Owner", "OWNER TODO", "P1", "GA4 Admin → Custom definitions → event-scoped. Fields already pushed to dataLayer."),
    ("Verify all events in GA4 Realtime after deploy + test booking",
     "Owner", "OWNER TODO", "P0", "form_submitted → lead_verifided → thankyou_conversion → demo_booked."),
]

META = [
    ("Meta Pixel base live via GTM (Pixel ID 2862017797322752)",
     "Code/Owner", "DONE", "P0", "Through container GTM-K5D84Z3L."),
    ("FB - Book Appointment tag (Standard event Schedule) on calendly Trigger",
     "Owner", "DONE", "P0", "Created by owner."),
    ("Unpause FB - OTP Verified tag (trigger lead_verifided)",
     "Owner", "DONE", "P0", "Owner unpaused."),
    ("Manual Advanced Matching on FB - Book demo (6 params)",
     "Owner", "DONE", "P0", "Mapped first_name/last_name/phone/email/external_id(=phone)/city_name from dataLayer."),
    ("Replicate the 6-row Advanced Matching block on FB - OTP Verified + Schedule tags",
     "Owner", "OWNER TODO", "P0", "Code already feeds all 6 fields on every lead/booking event."),
    ("Map conversion_value (+ currency INR) into Meta conversion tags",
     "Owner", "OWNER TODO", "P1", "lead_verified=₹500, demo_booked=₹2000."),
    ("Publish container + verify Event Match Quality (EMQ) in Events Manager Test Events",
     "Owner", "OWNER TODO", "P0", "After deploy — confirm AM raises match-rate."),
]

GADS = [
    ("Google Ads base / conversion linker live (AW-16740091756 / NtqdClejmOgaEOyOpq4-)",
     "Code/Owner", "DONE", "P0", "Through container GTM-K5D84Z3L."),
    ("GAds - Book Demo fires on lead_verifided (verified event)",
     "Owner", "DONE", "P0", "Earlier 'repoint to lead_verified' task is MOOT — already on the verified event."),
    ("Turn ON Enhanced Conversions for leads + map User-Provided Data (email/phone/first/last/city)",
     "Owner", "OWNER TODO", "P0", "BIGGEST lever. Raw fields already in dataLayer; GTM hashes automatically."),
    ("Create / import 'Book appointment' conversion from GA4 demo_booked key event",
     "Owner", "OWNER TODO", "P0", "Existing 'Book appointments' is Import-from-clicks (Zapier) — no tag/ID/label, so use GA4 key-event import. Set Primary."),
    ("Map conversion_value into Ads conversion tags; set demo_booked as value-based secondary",
     "Owner", "OWNER TODO", "P1", "Confirm ₹2000 is right for a booked demo vs real avg deal value."),
    ("Build junk-exclusion / suppression audience where lead_quality = junk",
     "Owner", "OWNER TODO", "P2", "Stop paying to retarget bots / too-fast submits."),
    ("Turn OFF the Zapier appointment import once browser demo_booked conversion is confirmed",
     "Owner", "OWNER TODO", "P0", "Prevents double-counting in Google Ads. Do AFTER verifying the new conversion fires."),
]

DEPLOY = [
    ("Set REACT_APP_BACKEND_URL to the PRODUCTION backend URL before build",
     "Owner/Code", "OWNER TODO", "P0", "GATE 3 (functional) — currently preview URL; would break the demo form."),
    ("Keep REACT_APP_GTM_ID=GTM-K5D84Z3L in frontend/.env",
     "Code", "DONE", "P0", "Re-added after env recreation. GTM is host-gated; without it NO tags fire in prod."),
    (".gitignore no longer blocks .env files from deploying",
     "Code", "DONE", "P0", "Removed .env/.env.*/*.env ignore lines so backend/.env + frontend/.env ship."),
    ("deployment_agent readiness check",
     "Code", "DONE", "P0", "PASS. DB .to_list(1000/5000) caps deemed non-blocking admin patterns."),
    ("Click Deploy (Emergent) → site live on www.mygenie.online",
     "Owner", "OWNER TODO", "P0", "First time GTM actually loads (host-gated)."),
    ("Verify SSL green on https://www.mygenie.online",
     "Owner", "OWNER TODO", "P0", "After DNS flip."),
    ("Re-register Calendly webhook to the production external URL",
     "Code", "BLOCKED", "P0", "Awaiting deploy. POST /api/webhooks/calendly — needed so bookings hit backend + Freshsales flow."),
    ("Run GTM Preview after deploy — confirm every tag fires on its trigger",
     "Owner", "OWNER TODO", "P0", "form_submitted → lead_verifided → thankyou_conversion → demo_booked."),
    ("Do 1 live test booking end-to-end (form → OTP → Calendly)",
     "Owner", "OWNER TODO", "P0", "Drives the GA4 key-event + Ads conversion import steps."),
    ("Watch rollback triggers in first 1-24h (old URLs 404, canonical broken, blank render, 5xx)",
     "Owner", "OWNER TODO", "P0", "Rollback = repoint DNS back to old origin (fast, non-destructive)."),
    ("Delete leftover QA/test contacts from Freshsales",
     "Owner", "OWNER TODO", "P2", "e.g. 402209074834 / 402209074908 + 'Smoke Test'/'QA Tester'."),
]

# Summary tab content
SUMMARY = [
    ("SEO", "Owner+Code", "READY", "P0", "Build is SEO-ready; CONDITIONAL GO. Gates 1-4 must be live at DNS flip. 301 layer = single most important item."),
    ("GA4", "Owner+Code", "READY", "P0", "Tags built & published. Remaining = mark demo_booked key event, unpause OTP tag, map values/dimensions."),
    ("Meta", "Owner+Code", "READY", "P0", "Pixel + AM on Book demo done. Remaining = replicate AM on OTP/Schedule tags, map values, verify EMQ."),
    ("Google Ads", "Owner+Code", "READY", "P0", "Tags fire on correct events. Remaining = Enhanced Conversions, demo_booked conversion from GA4, kill Zapier."),
    ("Deployment", "Owner+Code", "READY", "P0", "deployment_agent PASS. Remaining = set prod backend URL, Deploy, Calendly webhook re-register, GTM Preview verify."),
]


def build():
    wb = Workbook()
    sub = f"MyGenie POS — generated {date.today().isoformat()} · container GTM-K5D84Z3L · GA4 G-KWHHFEZ5Q3 · Pixel 2862017797322752 · Ads AW-16740091756"

    ws0 = wb.active
    ws0.title = "Summary"
    style_sheet(ws0, "Pre-Production Checklist — Summary",
                "Status legend: DONE · OWNER TODO · BLOCKED · READY · OPTIONAL  |  Priority: P0 (blocker) · P1 · P2", SUMMARY)

    for name, rows in [("SEO", SEO), ("GA4", GA4), ("Meta", META),
                       ("Google Ads", GADS), ("Deployment", DEPLOY)]:
        ws = wb.create_sheet(name)
        style_sheet(ws, f"{name} Checklist", sub, rows)

    wb.save(OUT)
    print(f"Wrote {OUT}")
    print(f"Tabs: {wb.sheetnames}")


if __name__ == "__main__":
    build()
