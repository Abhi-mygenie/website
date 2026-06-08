"""Generate the MyGenie POS V2.4 Website handover & inputs tracker (Excel)."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.worksheet.datavalidation import DataValidation

GREEN = "18A84A"
DEEP = "10402A"
ORANGE = "F26822"
SAND = "F6F8F5"
LINE = "E6EAE6"
WHITE = "FFFFFF"

wb = openpyxl.Workbook()

thin = Side(style="thin", color=LINE)
border = Border(left=thin, right=thin, top=thin, bottom=thin)
hdr_font = Font(name="Calibri", bold=True, color=WHITE, size=11)
title_font = Font(name="Calibri", bold=True, color=DEEP, size=16)
sub_font = Font(name="Calibri", color="5B6B61", size=10, italic=True)
cell_font = Font(name="Calibri", size=10, color="14201A")
wrap = Alignment(wrap_text=True, vertical="top")
center = Alignment(horizontal="center", vertical="center")


def style_sheet(ws, title, subtitle, headers, rows, widths, status_col=None,
                priority_col=None):
    ws.sheet_view.showGridLines = False
    # Title
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=len(headers))
    c = ws.cell(row=1, column=1, value=title)
    c.font = title_font
    ws.row_dimensions[1].height = 26
    ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=len(headers))
    c2 = ws.cell(row=2, column=1, value=subtitle)
    c2.font = sub_font
    # Header row (row 4)
    hr = 4
    for j, h in enumerate(headers, start=1):
        cell = ws.cell(row=hr, column=j, value=h)
        cell.font = hdr_font
        cell.fill = PatternFill("solid", fgColor=GREEN)
        cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
        cell.border = border
    ws.row_dimensions[hr].height = 22
    # Data
    for i, row in enumerate(rows):
        r = hr + 1 + i
        fill = SAND if i % 2 == 0 else WHITE
        for j, val in enumerate(row, start=1):
            cell = ws.cell(row=r, column=j, value=val)
            cell.font = cell_font
            cell.alignment = wrap
            cell.border = border
            cell.fill = PatternFill("solid", fgColor=fill)
        ws.row_dimensions[r].height = 34
    # Column widths
    for j, w in enumerate(widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(j)].width = w
    # Freeze
    ws.freeze_panes = ws.cell(row=hr + 1, column=1)
    n = len(rows)
    # Status dropdown
    if status_col:
        dv = DataValidation(type="list", formula1='"Pending,In Progress,Provided,Done,N/A"', allow_blank=True)
        ws.add_data_validation(dv)
        col = openpyxl.utils.get_column_letter(status_col)
        dv.add(f"{col}{hr+1}:{col}{hr+n}")
    if priority_col:
        dv2 = DataValidation(type="list", formula1='"P0 - Blocker,P1 - Important,P2 - Nice to have"', allow_blank=True)
        ws.add_data_validation(dv2)
        col = openpyxl.utils.get_column_letter(priority_col)
        dv2.add(f"{col}{hr+1}:{col}{hr+n}")


# ---------------- SHEET 1: Inputs Needed From You ----------------
ws1 = wb.active
ws1.title = "1. Inputs Needed From You"
headers1 = ["#", "Category", "What we need from you", "Why / where it's used on the site",
            "Format", "Priority", "Status", "Your notes / link"]
rows1 = [
    [1, "Brand", "Official MyGenie logo files (light + dark versions)", "Replaces the temporary text logomark in navbar & footer", "SVG or transparent PNG", "P0 - Blocker", "Pending", ""],
    [2, "Brand", "Exact brand hex codes & font license (if any beyond Poppins)", "Confirm/lock current green #18A84A, orange #F26822, yellow #FFC529", "Doc / values", "P1 - Important", "Pending", ""],
    [3, "Pricing", "Real prices for 4 plans (Starter, Growth, Pro, Chain) - monthly & annual", "Pricing builder currently uses PLACEHOLDER numbers", "Excel/values", "P0 - Blocker", "Pending", ""],
    [4, "Pricing", "Real prices for 9 add-ons (Captain device, WhatsApp, Loyalty, Scan&Order, Delivery, Central Inv, Hotel billing, AI, Support)", "Add-on cart in pricing builder", "Excel/values", "P0 - Blocker", "Pending", ""],
    [5, "Pricing", "Confirm what's included in each plan (feature lists)", "Plan cards + 'Included/Free in plan' logic", "List per plan", "P1 - Important", "Pending", ""],
    [6, "Pricing", "Free trial? Setup fee? Money-back? Hardware cost?", "Pricing page reassurance + FAQ", "Text", "P2 - Nice to have", "Pending", ""],
    [7, "Proof", "Real aggregate stats: outlets powered, cities, orders processed, total saved, avg rating", "Homepage trust band + Success Stories stats (currently indicative)", "Values", "P0 - Blocker", "Pending", ""],
    [8, "Proof", "Approved client logos + names for the logo wall", "Homepage trust marquee + customer logos", "PNG/SVG logos + approval", "P1 - Important", "Pending", ""],
    [9, "Proof", "Which case studies can be named publicly (sign-off)", "Success Stories, sector & product proof sections", "Approval list", "P1 - Important", "Pending", ""],
    [10, "Media", "Feature product videos for the 5 product pages", "Drops into the 'See it in action' video slot (videoUrl)", "YouTube/Vimeo/MP4 links", "P1 - Important", "Pending", ""],
    [11, "Media", "Real product screenshots (dashboard, KDS, captain app, room checkout, scan-order)", "Replace stock images in heroes/sections", "PNG/JPG", "P1 - Important", "Pending", ""],
    [12, "AI", "Confirm which AI features are LIVE vs roadmap", "Keeps AI copy honest (menu import, insights, upsell, validations, reporting, segmentation)", "Yes/No per feature", "P0 - Blocker", "Pending", ""],
    [13, "Product", "Hotel/PMS detail: standalone room billing or PMS integration? what 'single bill' covers", "Hotels & Resorts page accuracy", "Text", "P1 - Important", "Pending", ""],
    [14, "Product", "Confirm integrations list (Swiggy, Zomato, Magicpin, Rapido, Dunzo, payments, accounting)", "Cloud Kitchen page + future integrations page", "List + status", "P1 - Important", "Pending", ""],
    [15, "Payments", "Razorpay decision for live 'Buy Online' + keys if yes", "Pricing builder checkout currently captures order only (no charge)", "Yes/No + Key ID/Secret", "P1 - Important", "Pending", ""],
    [16, "Leads", "Where should demo & quote leads go? (CRM / email / Google Sheet / API)", "Currently saved to MyGenie DB only; needs routing to your sales", "CRM name + API/email", "P0 - Blocker", "Pending", ""],
    [17, "Contact", "Business phone, WhatsApp number, support email, address", "Footer, contact, 'Talk to us on WhatsApp' option", "Text", "P1 - Important", "Pending", ""],
    [18, "Contact", "Social media links (Instagram, LinkedIn, etc.)", "Footer social icons", "URLs", "P2 - Nice to have", "Pending", ""],
    [19, "Legal", "Privacy Policy & Terms content", "Footer legal pages (required before go-live)", "Text/Docs", "P1 - Important", "Pending", ""],
    [20, "GTM", "Priority geographies & languages (India only? GCC? multi-language?)", "Copy tone, currency, future localization", "Text", "P2 - Nice to have", "Pending", ""],
    [21, "GTM", "Sales reps available per sector (for sector-specific demo routing)", "Sector-aware demo follow-up", "Names/roles", "P2 - Nice to have", "Pending", ""],
    [22, "Deploy", "Domain & DNS access to go live (e.g., mygenie.online or a subdomain)", "Production deployment", "Domain + DNS access", "P1 - Important", "Pending", ""],
]
widths1 = [5, 13, 42, 46, 22, 18, 15, 30]
style_sheet(ws1, "MyGenie POS V2.4 - Inputs Needed From You",
            "Fill the Status & Notes columns as you gather each item. P0 = blocks go-live, P1 = important, P2 = nice to have.",
            headers1, rows1, widths1, status_col=7, priority_col=6)

# ---------------- SHEET 2: Pages & Status ----------------
ws2 = wb.create_sheet("2. Pages Built & Status")
headers2 = ["#", "Page / Feature", "URL (preview)", "What it does", "Build status", "Uses placeholder?"]
BASE = "https://ai-inventory-hub-6.preview.emergentagent.com"
rows2 = [
    [1, "Homepage", BASE + "/", "12 conversion sections: hero, pains, before/after, pillars, sectors, modules, AI band, proof, demo", "Built & tested", "Trust numbers, logo"],
    [2, "Pricing Builder", BASE + "/pricing", "Recommendation quiz + 4 plans + add-on cart + upsell/cross-sell + Buy/Demo", "Built & tested", "Prices, payment"],
    [3, "Sector: Restaurants", BASE + "/solutions/restaurants", "Pain-led sector landing + proof + FAQ + sector demo", "Built & tested", "Images"],
    [4, "Sector: Cafes", BASE + "/solutions/cafes", "Sector landing page", "Built & tested", "Images"],
    [5, "Sector: QSR", BASE + "/solutions/qsr", "Sector landing page", "Built & tested", "Images"],
    [6, "Sector: Cloud Kitchens", BASE + "/solutions/cloud-kitchens", "Sector landing page", "Built & tested", "Images"],
    [7, "Sector: Hotels & Resorts", BASE + "/solutions/hotels-resorts", "Sector landing page", "Built & tested", "Images"],
    [8, "Sector: Food Courts", BASE + "/solutions/food-courts", "Sector landing page", "Built & tested", "Images"],
    [9, "Sector: Canteens & Mess", BASE + "/solutions/canteens", "Sector landing page", "Built & tested", "Images"],
    [10, "Sector: Chains & Franchises", BASE + "/solutions/chains", "Sector landing page", "Built & tested", "Images"],
    [11, "Product: Sell & Serve", BASE + "/product/sell-serve", "Module bucket page + video slot", "Built & tested", "Video, images"],
    [12, "Product: Run the Property", BASE + "/product/run-property", "Module bucket page + video slot", "Built & tested", "Video, images"],
    [13, "Product: Bring Customers Back", BASE + "/product/customers", "Module bucket page + video slot", "Built & tested", "Video, images"],
    [14, "Product: Protect Your Profit", BASE + "/product/protect-profit", "Module bucket page + video slot", "Built & tested", "Video, images"],
    [15, "Product: See Everything", BASE + "/product/see-everything", "Module bucket page + video slot", "Built & tested", "Video, images"],
    [16, "Success Stories", BASE + "/customers", "15 filterable case studies + stats band", "Built & tested", "Aggregate stats"],
    [17, "ROI Calculator", BASE + "/roi", "Live savings estimate + lead capture", "Built & tested", "Assumption rates"],
    [18, "Help & FAQ", BASE + "/resources", "9-question FAQ accordion", "Built & tested", "No"],
    [19, "Practical AI page", "(not built)", "Deferred per your direction - you'll share feature videos", "Deferred", "-"],
]
widths2 = [5, 30, 52, 46, 16, 18]
style_sheet(ws2, "MyGenie POS V2.4 - Pages Built & Status",
            "17 live routes. 'Uses placeholder?' shows what gets swapped once you provide the matching input from Sheet 1.",
            headers2, rows2, widths2)

# ---------------- SHEET 3: Integrations & Deployment ----------------
ws3 = wb.create_sheet("3. Integrations & Deploy")
headers3 = ["#", "Item", "Current state", "Action needed", "Owner", "Status"]
rows3 = [
    [1, "Lead capture (demo form)", "Saves to MyGenie database (/api/demo-request)", "Decide CRM/email destination & connect", "You + Dev", "Pending"],
    [2, "Quote capture (pricing builder)", "Saves to database (/api/quote)", "Route to sales / CRM", "You + Dev", "Pending"],
    [3, "Online payment (Buy Online)", "Captures order only - does NOT charge", "Add Razorpay (provide keys) to enable real payment", "You", "Pending"],
    [4, "WhatsApp 'Talk to us'", "Not added", "Provide WhatsApp business number to add click-to-chat", "You", "Pending"],
    [5, "Swiggy/Zomato/Magicpin", "Mentioned in copy only", "Confirm live integrations; build integrations page later", "You", "Pending"],
    [6, "Analytics (GA4 / Meta Pixel)", "Not added", "Provide IDs to track conversions", "You", "Pending"],
    [7, "Domain & deployment", "Running on preview URL", "Choose domain; deploy to production", "You + Dev", "Pending"],
    [8, "Product videos embed", "Placeholder slot ready (videoUrl=null)", "Provide video links to auto-embed", "You", "Pending"],
]
widths3 = [5, 30, 42, 46, 16, 15]
style_sheet(ws3, "MyGenie POS V2.4 - Integrations & Deployment", "Technical hookups pending your inputs/decisions.",
            headers3, rows3, widths3, status_col=6)

# ---------------- SHEET 4: Next Steps Roadmap ----------------
ws4 = wb.create_sheet("4. Next Steps Roadmap")
headers4 = ["#", "Step", "Description", "Depends on (from Sheet 1)", "Priority", "Status"]
rows4 = [
    [1, "Swap in real prices", "Replace placeholder prices in pricing builder", "Items 3, 4, 5", "P0 - Blocker", "Pending"],
    [2, "Swap in brand assets", "Add official logo + confirm colors", "Items 1, 2", "P0 - Blocker", "Pending"],
    [3, "Swap in real proof", "Real stats + approved logos + named case studies", "Items 7, 8, 9", "P0 - Blocker", "Pending"],
    [4, "Add product videos", "Embed feature videos on 5 product pages", "Item 10", "P1 - Important", "Pending"],
    [5, "Connect leads to CRM", "Route demo + quote leads to your sales system", "Item 16", "P0 - Blocker", "Pending"],
    [6, "Enable online payment", "Wire Razorpay for live Buy Online", "Item 15", "P1 - Important", "Pending"],
    [7, "Add legal + contact", "Privacy/Terms pages + contact details + socials", "Items 17, 18, 19", "P1 - Important", "Pending"],
    [8, "Add analytics", "GA4 / Meta Pixel for conversion tracking", "Item (analytics)", "P1 - Important", "Pending"],
    [9, "Build Practical AI page", "When feature videos are ready", "Items 10, 12", "P2 - Nice to have", "Pending"],
    [10, "SEO & content", "Blog/guide clusters, integrations page, comparison pages", "-", "P2 - Nice to have", "Pending"],
    [11, "Deploy to production", "Go live on your domain", "Items 22 + all P0", "P1 - Important", "Pending"],
    [12, "ROI report email + CRM", "Email PDF savings report + pipe to CRM (suggested enhancement)", "Item 16", "P2 - Nice to have", "Pending"],
]
widths4 = [5, 28, 50, 28, 18, 15]
style_sheet(ws4, "MyGenie POS V2.4 - Next Steps Roadmap", "Recommended order. Complete P0 items to be go-live ready.",
            headers4, rows4, widths4, status_col=6, priority_col=5)

# ---------------- SHEET 5: Image Mapping Plan ----------------
ws5 = wb.create_sheet("5. Image Mapping Plan")
headers5 = ["#", "Owned asset (from your live site)", "What it is", "Best placement in V2.4 site",
            "Replaces (current stock)", "Priority", "Status", "Notes / decision"]
rows5 = [
    [1, "logo.svg (+ app/social icons)", "Official MyGenie logo", "Navbar + Footer", "Temporary text logomark", "P0 - Blocker", "Pending", "Use as-is"],
    [2, "banner.png (776x637)", "Hero illustration", "Homepage hero visual", "Pexels 4339955 (owner photo)", "P1 - Important", "Pending", "Confirm suits new hero; else use product screenshot"],
    [3, "vidban.jpg (1600x900) + YouTube video", "Video banner / overview", "Product page 'See it in action' slots + homepage", "Empty video slot / stock hero", "P1 - Important", "Pending", "Provide YouTube/Vimeo URL to embed"],
    [4, "feature1-5.png (300x300)", "5 feature illustrations", "5 Product page heroes + module overview", "Pexels product heroes (5 photos)", "P1 - Important", "Pending", "VERIFY what each depicts to map to correct bucket; check resolution for hero use"],
    [5, "i1-i9.svg", "Feature line icons", "Module/feature icons (optional)", "lucide icons (already license-safe)", "P2 - Nice to have", "Pending", "Optional - lucide is fine to keep"],
    [6, "8 client logos (Hyatt, Bamboo Yoga, Palm Forest, Baba's Italy, Love Bites, Mill Bakery, Wild Berry, Drishti Yoga)", "Brand/client logos 250x250", "Homepage trust marquee + sector 'trusted by' strips", "Placeholder text logo names", "P1 - Important", "Pending", "Confirm display rights/approval per client"],
    [7, "11 case-study images (lovebites, ubuntu, sushi, mill, palmforest, pavandpages, terra, lafetta, beanmeup, luxevista, bambooyoga)", "Case study photos/logos 250x250", "Success Stories cards + sector/product proof", "Text-only cards (no image now)", "P1 - Important", "Pending", "Map each to its matching testimonial client"],
    [8, "form-img.jpg (738x765)", "Demo form side image", "Demo sections (home/sector/product)", "(decorative, none)", "P2 - Nice to have", "Pending", "Optional visual"],
    [9, "apple.svg, google.svg", "App Store / Play badges", "Footer 'download app'", "Not present", "P2 - Nice to have", "Pending", "Add app store links"],
    [10, "youtube.png, facebook.png, mail.svg, phone1.png", "Social + contact icons", "Footer social + contact row", "Placeholder", "P1 - Important", "Pending", "Provide live social URLs (Sheet 1, item 18)"],
    [11, "qut.png", "Quote mark graphic", "Testimonial quote decoration", "lucide Quote icon", "P2 - Nice to have", "Pending", "Optional"],
    [12, "(GAP) No 8 distinct sector photos", "Sector hero imagery missing", "8 Sector page heroes", "Pexels sector photos (8)", "P0 - Blocker", "Done", "RESOLVED: using brand-graphic treatment (icon + metric on deep-green panel). No photos needed."],
    [13, "feature1-5.png mapping", "Used as product-page hero illustrations (for now)", "5 Product page heroes", "Pexels product photos", "P1 - Important", "In Progress", "VERIFY each illustration matches its bucket (1=Sell&Serve, 2=Run Property, 3=Customers, 4=Protect Profit, 5=See Everything). Replace if content mismatches or low-res."],
    [14, "Case-study image accuracy", "11 case images now on Success Stories cards", "Success Stories + proof", "Text-only cards", "P1 - Important", "In Progress", "Confirm each image matches the right client; some clients have no image (initials shown): Rhino, Matryyoshka, Kates, Antonios."],
]
widths5 = [5, 40, 26, 38, 30, 18, 14, 40]
style_sheet(ws5, "MyGenie POS V2.4 - Image Mapping Plan (replace ALL stock with owned assets)",
            "Strategic plan only - no changes made yet. 43 owned images inventoried from mygenie.online. Confirm decisions, then we execute the swap.",
            headers5, rows5, widths5, status_col=7, priority_col=6)

out = "/app/MyGenie_V2.4_Handover_Tracker.xlsx"
wb.save(out)
print("Saved", out)
